import os
import re
import sys
from datetime import date

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(ROOT, '.env'))
except ImportError:
    pass

from db.connection import get_connection
from db.queries.schedule import (
    clear_schedule_for_upload,
    insert_upload,
    insert_lesson,
    get_or_create_schedule_teacher,
    get_or_create_classroom,
)
from db.queries.groups import get_or_create_group

BASE_DIR = os.path.dirname(__file__)

if len(sys.argv) > 1:
    FILE_NAME = sys.argv[1]
else:
    FILE_NAME = os.path.join(BASE_DIR, 'uploads', 'schedule.xlsx')

effective_from = None
source = 'manual'
vk_post_id = None
for i, arg in enumerate(sys.argv):
    if arg == '--effective-from' and i + 1 < len(sys.argv):
        parts = sys.argv[i + 1].split('.')
        if len(parts) == 3:
            effective_from = date(int(parts[2]), int(parts[1]), int(parts[0]))
    elif arg == '--source' and i + 1 < len(sys.argv):
        source = sys.argv[i + 1]
    elif arg == '--vk-post-id' and i + 1 < len(sys.argv):
        vk_post_id = int(sys.argv[i + 1])

if not os.path.exists(FILE_NAME):
    print('Excel файл не найден:', FILE_NAME)
    sys.exit(1)


def split_time(text):
    if '-' in text:
        arr = text.split('-')
        return arr[0].strip(), arr[1].strip()
    return '', ''


wb = load_workbook(FILE_NAME, data_only=True)
ws = wb.active

merged = {}
for rng in ws.merged_cells.ranges:
    min_col, min_row, max_col, max_row = rng.bounds
    val = ws.cell(min_row, min_col).value
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            merged[(r, c)] = val


def get_val(row, col):
    val = ws.cell(row, col).value
    if val is None:
        val = merged.get((row, col))
    if val is None:
        return ''
    return str(val).strip()


conn = get_connection()
try:
    clear_schedule_for_upload(conn)
    upload_id = insert_upload(
        conn,
        os.path.basename(FILE_NAME),
        effective_from=effective_from,
        source=source,
        vk_post_id=vk_post_id,
    )

    GROUP_ROW = 19
    HEADER_ROW = 20
    group_columns = {}
    for col in range(1, ws.max_column + 1):
        header = get_val(HEADER_ROW, col)
        if header == 'Дисциплина,модуль':
            group_text = get_val(GROUP_ROW, col)
            names = re.findall(r'[А-Яа-яA-Za-z0-9\-]+-\d+-\d+-\d+', group_text)
            for name in names:
                group_columns[name] = col

    print('Найдено групп:', len(group_columns))

    for group_name, col in group_columns.items():
        print('Обработка:', group_name)
        group_id = get_or_create_group(conn, group_name)
        current_day = ''
        lesson_number = 0
        for row in range(21, ws.max_row + 1):
            day = get_val(row, 7)
            time_text = get_val(row, 8)
            if day:
                current_day = day.strip()
                lesson_number = 0
            subject = get_val(row, col)
            lesson_type = get_val(row, col + 1)
            teacher = get_val(row, col + 2)
            room = get_val(row, col + 3)
            if subject:
                lesson_number += 1
                teacher_id = get_or_create_schedule_teacher(conn, teacher)
                classroom_id = get_or_create_classroom(conn, room)
                time_start, time_end = split_time(time_text)
                insert_lesson(
                    conn, upload_id, group_id, teacher_id, classroom_id,
                    current_day, lesson_number, time_start, time_end, subject, lesson_type,
                )
    conn.commit()
finally:
    conn.close()

print('Расписание успешно загружено в PostgreSQL')
